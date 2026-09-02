"""
=====================================================
  Order Manager — Excel এ অর্ডার সেভ করে
=====================================================
"""

import openpyxl
import os
from datetime import datetime

ORDERS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "orders.xlsx")

HEADERS = ["তারিখ", "নাম", "মোবাইল", "ঠিকানা", "পণ্য", "রঙ/সাইজ", "পেমেন্ট", "স্ট্যাটাস"]


def _get_workbook():
    if os.path.exists(ORDERS_FILE):
        wb = openpyxl.load_workbook(ORDERS_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "অর্ডার লিস্ট"

        # Header styling
        ws.append(HEADERS)
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[1]:
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Column widths
        widths = [18, 18, 16, 30, 25, 20, 22, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    return wb, ws


def save_order(name, mobile, address, product, color_size="", payment="ক্যাশ অন ডেলিভারি"):
    """Excel এ নতুন অর্ডার সেভ করে"""
    try:
        wb, ws = _get_workbook()
        ws.append([
            datetime.now().strftime("%d/%m/%Y %H:%M"),
            name,
            mobile,
            address,
            product or "অনির্ধারিত",
            color_size or "-",
            payment,
            "নতুন"
        ])

        # Row styling — alternate colors
        row = ws.max_row
        from openpyxl.styles import PatternFill, Alignment
        fill_color = "DEEAF1" if row % 2 == 0 else "FFFFFF"
        fill = PatternFill("solid", fgColor=fill_color)
        for cell in ws[row]:
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        ws.row_dimensions[row].height = 20
        wb.save(ORDERS_FILE)
        print(f"[EXCEL] Order saved: {name} | {mobile} | {payment}")
        return True
    except Exception as e:
        print(f"[EXCEL ERROR] {e}")
        return False


def get_all_orders():
    """সব অর্ডার return করে (admin panel এর জন্য)"""
    try:
        if not os.path.exists(ORDERS_FILE):
            return []
        wb = openpyxl.load_workbook(ORDERS_FILE)
        ws = wb.active
        orders = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(c for c in row):
                orders.append(row)
        return orders
    except Exception as e:
        print(f"[EXCEL READ ERROR] {e}")
        return []
